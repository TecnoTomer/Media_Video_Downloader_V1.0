import tkinter as tk
from tkinter import ttk
from PIL import Image, ImageTk
from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import os.path
import os
import webbrowser
import json
import pandas as pd
from tkinter import filedialog
from pytube import YouTube
import time
import pyperclip

import Variables
import Youtube_Funciones
import Home_Funciones
import Alertas

def selecion_boton(self):
    self.bind("<Enter>", lambda event: self.config(bg=Variables.c_selecion))
    self.bind("<Leave>", lambda event: self.config(bg=Variables.c_barras))

def seleccionar_label(label, imagen_original, imagen_hover):
    def cambiar_imagen(event):
        if event.type == tk.EventType.Enter:
            label.config(image=imagen_hover)
        elif event.type == tk.EventType.Leave:
            label.config(image=imagen_original)

    label.bind("<Enter>", cambiar_imagen)
    label.bind("<Leave>", cambiar_imagen)

def cambio_ventana(ventana_factory, frame):
    for widget in frame.winfo_children():
        widget.destroy()

    ventana = ventana_factory(frame)
    ventana.pack(side="top", fill="both", expand=True)

def logos(logo, ancho, alto, x, y):
    image = Image.open(logo).convert("RGBA")
    image = image.resize((ancho, alto), Image.ANTIALIAS)

    fondo_transparente = Image.new("RGBA", image.size, (255, 255, 255, 0))
    image_con_fondo_transparente = Image.alpha_composite(fondo_transparente, image)

    photo = ImageTk.PhotoImage(image_con_fondo_transparente)

    label_image = tk.Label(Youtube_Funciones.frame1_botones, image=photo, bg=Variables.Y_backgroun)
    label_image.image = photo 
    label_image.place(x=x, y=y)

    return label_image

def search_boton(logo, hover, ancho, alto, x, y):
    image_original = Image.open(logo).convert("RGBA")
    image_original = image_original.resize((ancho, alto), Image.ANTIALIAS)
    photo_original = ImageTk.PhotoImage(image_original)
    label_image = tk.Label(Youtube_Funciones.frame1_botones, image=photo_original, bg=Variables.Y_backgroun)
    label_image.place(x=x, y=y)
    imagen_hover = Image.open(hover)
    imagen_hover = imagen_hover.resize((ancho, alto), Image.ANTIALIAS)
    photo_hover = ImageTk.PhotoImage(imagen_hover)

    seleccionar_label(label_image, photo_original, photo_hover)

    return label_image

def Barra_busqueda(parent, width, x, y): #barra busqueda
    placeholder = "Introduce URL del video o playList"

    def borrar_placeholder(event):
        if entry.get() == placeholder:
            entry.delete(0, tk.END)
            entry.config(fg='white')  # Cambiar el color del texto al predeterminado

    def restaurar_placeholder(event):
        if not entry.get():
            entry.insert(0, placeholder)
            entry.config(fg='grey')  # Cambiar el color del texto a gris

    font_style = (Variables.poppins_negrita, 13)  # Tipo de letra y tamaño del texto
    entry = tk.Entry(parent, width=width, fg='grey', bg=Variables.Y_backgroun, font=font_style, borderwidth=0)
    entry.insert(0, placeholder)
    entry.bind("<FocusIn>", borrar_placeholder)
    entry.bind("<FocusOut>", restaurar_placeholder)
    entry.place(x=x, y=y)  # Establecer la posición manualmente

    return entry

def crear_excel_lista(datos, nombre_archivo):
    try:
        # Verificar si el archivo Excel existe
        if os.path.exists(nombre_archivo):
            # Cargar el libro de Excel existente
            wb = load_workbook(nombre_archivo)
            ws = wb.active
            # Obtener el número de la próxima fila vacía
            fila = ws.max_row + 1
        else:
            # Crear un nuevo libro de Excel
            wb = Workbook()
            # Seleccionar la hoja activa (por defecto la primera)
            ws = wb.active
            # Escribir los encabezados
            ws['A1'] = 'Título'
            ws['B1'] = 'Duración'
            ws['C1'] = 'Estado'
            ws['D1'] = 'Link'
            fila = 2  # Comenzar desde la fila 2 después de los encabezados
        
        # Escribir los datos en las celdas específicas
        if isinstance(datos, list):
            for fila_datos in datos:
                ws.cell(row=fila, column=1, value=fila_datos[0])  # Título
                ws.cell(row=fila, column=2, value=fila_datos[1])  # Duración
                ws.cell(row=fila, column=3, value=fila_datos[2])  # Estado
                ws.cell(row=fila, column=4, value=fila_datos[3])  # Link
                fila += 1  # Incrementar la fila para la próxima iteración

        # Guardar el libro de Excel
        wb.save(nombre_archivo)
        
    except Exception as e:
        pass

def crear_excel_video(datos, nombre_archivo):
    try:
        # Verificar si el archivo Excel existe
        if os.path.exists(nombre_archivo):
            # Cargar el libro de Excel existente
            wb = load_workbook(nombre_archivo)
            ws = wb.active
            # Obtener el número de la próxima fila vacía
            fila = ws.max_row + 1
        else:
            # Crear un nuevo libro de Excel
            wb = Workbook()
            # Seleccionar la hoja activa (por defecto la primera)
            ws = wb.active
            # Escribir los encabezados
            ws['A1'] = 'Título'
            ws['B1'] = 'Duración'
            ws['C1'] = 'Estado'
            ws['D1'] = 'Link'
            fila = 2  # Comenzar desde la fila 2 después de los encabezados
        
        # Escribir los datos en las celdas específicas
        ws.cell(row=fila, column=1, value=datos[0])  # Título
        ws.cell(row=fila, column=2, value=datos[1])  # Duración
        ws.cell(row=fila, column=3, value=datos[2])  # Estado
        ws.cell(row=fila, column=4, value=datos[3])  # Link
        
        # Guardar el libro de Excel
        wb.save(nombre_archivo)
        
    except Exception as e:
        pass

def leer_datos_excel(nombre_archivo):
    try:
        # Cargar el libro de Excel
        wb = load_workbook(nombre_archivo)
        # Seleccionar la hoja activa (por defecto la primera)
        ws = wb.active
        # Leer los datos desde la fila 2 en adelante
        datos = []
        for fila in ws.iter_rows(min_row=2, values_only=True):
            datos.append(fila)
        return datos
    except Exception as e:
        print("Error al leer el archivo Excel:", e)
        return []

def crear_treeview(frame, nombre_archivo):
    global treeview
    # Crear el treeview
    treeview = ttk.Treeview(frame)

    # Configurar las columnas
    treeview["columns"] = ("Título", "Duración", "Estado")
    treeview.column("#0", width=0, stretch=tk.NO)  # Columna oculta
    treeview.column("Título", anchor=tk.W, width=450)
    treeview.column("Duración", anchor=tk.CENTER, width=100)
    treeview.column("Estado", anchor=tk.CENTER, width=100)

    # Encabezados de las columnas
    treeview.heading("#0", text="", anchor=tk.W)
    treeview.heading("Título", text="Título", anchor=tk.W)
    treeview.heading("Duración", text="Duración", anchor=tk.CENTER)
    treeview.heading("Estado", text="Estado", anchor=tk.CENTER)

    # Colocar el treeview en el frame
    treeview.pack(expand=True, fill="both")

    # Si el archivo Excel existe, leer los datos y actualizar el treeview
    if os.path.isfile(nombre_archivo):
        actualizar_treeview(treeview, nombre_archivo)

    return treeview

# Función para actualizar el treeview con datos desde el archivo Excel
def actualizar_treeview(treeview, nombre_archivo):
    # Leer los datos desde el archivo Excel
    datos = leer_datos_excel(nombre_archivo)
    
    # Borrar todos los elementos actuales del treeview
    treeview.delete(*treeview.get_children())

    # Configurar el ancho y la alineación de la columna "Estado"
    treeview.column("Estado", anchor=tk.CENTER, width=100)

    # Insertar los nuevos datos en el treeview
    for index, fila in enumerate(datos, start=1):
        # Insertar la fila en el treeview
        iid = treeview.insert("", index, text=index, values=fila)

        # Obtener el valor de la columna "Estado" de la fila actual
        estado = fila[2]  # Índice 2 corresponde a la columna "Estado" (0-indexed)

        # Verificar si el estado contiene "Terminado" y cambiar el color de fondo de la celda correspondiente
        if "Terminado" in estado:
            treeview.item(iid, tags=("terminado",))

    # Configurar el tag para cambiar el color de fondo
    treeview.tag_configure("terminado", background="green")

def abrir_link(consulta):
    try:
        consulta = consulta.replace(" ", "+")  # Reemplazar espacios con '+'
        url = f"https://www.youtube.com/results?search_query={consulta}"
        webbrowser.open(url)
    except Exception as e:
        print("Error al abrir el enlace:", e)

def clear(archivo):
    if os.path.exists(archivo):
        wb = load_workbook(archivo)
        ws = wb.active
        # Eliminar todas las filas excepto la primera (los encabezados)
        if ws.max_row > 1:
            ws.delete_rows(2, ws.max_row)
        # Guardar los cambios en el archivo Excel
        wb.save(archivo)
        actualizar_treeview(Youtube_Funciones.panel, Variables.data_excel)
    else:
        pass

def descargar(enlace, formato, resolucion):
    try:
        # Crear objeto YouTube para el video
        yt = YouTube(enlace)

        # Abrir el cuadro de diálogo para seleccionar la carpeta de destino
        carpeta_destino = filedialog.askdirectory()

        if carpeta_destino:
            if formato == "Mp4":
                video_stream = yt.streams.filter(file_extension="mp4", progressive=True, res=resolucion).first()
                if video_stream:
                    video_stream.download(output_path=carpeta_destino)
                    actualizar_estado(Variables.Terminado)
                    Msg = f"Descarga completa, verifique la ruta {carpeta_destino}"
                    Alertas.alerta_ok(Variables.titulo, "Done!", Msg)
                else:
                    Msg = f"La calidad selecionada {resolucion}, no esta disponible para este video, selecione una menor e intente de nuevo"
                    Alertas.alerta_Amarilla(Variables.titulo, "Error calidad", Msg)
            else:
                # Descargar audio en formato MP3
                audio_stream = yt.streams.filter(only_audio=True).first()
                if audio_stream:
                    audio_filename = audio_stream.default_filename
                    audio_stream.download(output_path=carpeta_destino)
                    # Cambiar la extensión del archivo a .mp3
                    audio_mp4_path = os.path.join(carpeta_destino, audio_filename)
                    audio_mp3_path = os.path.splitext(audio_mp4_path)[0] + ".mp3"
                    # Renombrar el archivo con la extensión .mp3
                    os.rename(audio_mp4_path, audio_mp3_path)

                    actualizar_estado(Variables.Terminado)

                    Msg = f"Descarga completa, verifique la ruta {carpeta_destino}"
                    Alertas.alerta_ok(Variables.titulo, "Done!", Msg)
        else:
            Msg = "No se seleccionó ninguna carpeta de destino."
            Alertas.alerta_Amarilla(Variables.titulo, "Alerta", Msg)
            pass

    except Exception as e:
            Msg = f"Ocurrio un error durante la descargar: {e}"
            Alertas.alerta_Amarilla(Variables.titulo, "Error imprevisto", Msg)

def actualizar_estado(nuevo_valor):
    try:
        # Leer el archivo Excel
        df = pd.read_excel(Variables.data_excel)
        fila = df[df['Link'] == Variables.enlace] #buscar columna Link y comparar valor de Variable.enlace

        if not fila.empty:
            # Obtener el índice de la fila encontrada
            indice_fila = fila.index[0]
            df.at[indice_fila, 'Estado'] = nuevo_valor
            
            # Guardar los cambios en el archivo Excel
            df.to_excel(Variables.data_excel, index=False)
            actualizar_treeview(Youtube_Funciones.panel, Variables.data_excel)
    except Exception as e:
        pass

def obtener_enlace_seleccionado():
    # Obtener el elemento seleccionado en el Treeview
    formato = Youtube_Funciones.formato_list.get()
    resolucion = Youtube_Funciones.calidades_list.get()
    
    # Verificar si el combobox formato está vacío
    if not formato:
        Mgs = "Debe seleccionar un formato de descarga Mp3 (No requiere resolucion) o Mp4 (Si requiere resolucion)."
        Alertas.alerta_Amarilla(Variables.titulo, "Campo formato sin seleccion", Mgs)
        return

    if formato == 'Mp4':
        if not resolucion:
            Mgs = "Formato Mp4, elija una resolucion."
            Alertas.alerta_Amarilla(Variables.titulo, "Campo resolucion sin seleccion", Mgs)
            return
        else:
            pass

    if formato == 'Mp3':
        if not resolucion:
            pass

    seleccion = treeview.selection()
    if seleccion:
        # Obtener el nombre del elemento seleccionado
        nombre_seleccionado = treeview.item(seleccion)['values'][0]

        # Leer el archivo Excel para buscar el enlace asociado al nombre
        try:
            df = pd.read_excel(Variables.data_excel)  # Leer el archivo Excel
            fila = df[df['Título'] == nombre_seleccionado]  # Buscar el nombre en la columna 'Título'
            if not fila.empty:
                Variables.enlace = fila.iloc[0]['Link']  # Obtener el enlace de la fila encontrada
                estado = fila.iloc[0]['Estado']
                name = fila.iloc[0]['Título']
                if "Pendiente" in estado:
                    descargar(Variables.enlace, formato, resolucion)
                else:
                    Msg = f"Ya se ha descargado {name}, ¿desea cambiar el estado?"
                    res = Alertas.alerta_aceptar(Variables.titulo, "Aviso", Msg)
                    if res == 1:
                        actualizar_estado(Variables.Pendiente)
                        Msg = "¿Desea continuar con la descarga?"
                        res = Alertas.alerta_aceptar_sin("Aviso", Msg)
                        if res == 1:
                            descargar(Variables.enlace, formato, resolucion)
                        else: pass
                    else: pass
            else:
                Mgs = "Error inprevisto, verifique que tenga algun elemento seleccionado, si no funciona cierre el programa y vuelva a intentar."
                Alertas.alerta_Amarilla(Variables.titulo, "Aviso", Mgs)
        except FileNotFoundError:
            Mgs = "Al parecer no se ha realizado una busqueda aun, por favor ingrese el URL de video o PlayList y vuelva a intentar."
            Alertas.alerta_Amarilla(Variables.titulo, "Aviso", Mgs)
    else:
        Mgs = "No se ha seleccionado ningún elemento de la lista, seleciona uno y vuelve a intentar."
        Alertas.alerta_Amarilla(Variables.titulo, "Aviso", Mgs)

#LISTAS DESCARGAR FUNCIONES
def actualizar_estado_lista(nuevo_valor, enlace):
    try:
        # Leer el archivo Excel
        df = pd.read_excel(Variables.data_excel)
        fila = df[df['Link'] == enlace]

        if not fila.empty:
            # Obtener el índice de la fila encontrada
            indice_fila = fila.index[0]
            df.at[indice_fila, 'Estado'] = nuevo_valor
            
            # Guardar los cambios en el archivo Excel
            df.to_excel(Variables.data_excel, index=False)
            actualizar_treeview(Youtube_Funciones.panel, Variables.data_excel)
    except Exception as e:
        pass

def descargar_todo():
    # Obtener el elemento seleccionado en el Treeview
    formato = Youtube_Funciones.formato_list.get()
    resolucion = Youtube_Funciones.calidades_list.get()
    
    # Verificar si el combobox formato está vacío
    if not formato:
        Mgs = "Debe seleccionar un formato de descarga Mp3 (No requiere resolucion) o Mp4 (Si requiere resolucion)."
        Alertas.alerta_Amarilla(Variables.titulo, "Campo formato sin seleccion", Mgs)
        return

    if formato == 'Mp4':
        if not resolucion:
            Mgs = "Formato Mp4, elija una resolucion."
            Alertas.alerta_Amarilla(Variables.titulo, "Campo resolucion sin seleccion", Mgs)
            return
        else:
            pass

    if formato == 'Mp3':
        if not resolucion:
            pass

    # Obtener la ruta de la carpeta de destino
    carpeta_destino = filedialog.askdirectory()

    # Verificar si se seleccionó una carpeta
    if carpeta_destino:
        # Crear una carpeta "Lista" dentro de la carpeta de destino
        nueva_carpeta = os.path.join(carpeta_destino, "Lista")
        os.makedirs(nueva_carpeta, exist_ok=True)  # La carpeta se crea si no existe

        try:
            # Leer el archivo Excel
            df = pd.read_excel(Variables.data_excel)

            # Iterar sobre las filas del archivo Excel comenzando desde Variables.fila
            for i in range(Variables.fila, len(df)):
                # Obtener los datos de la fila actual
                fila = df.iloc[i]

                # Obtener el enlace, formato y resolución de la fila
                enlace = fila['Link']
                estado = fila['Estado']
                name = fila['Título']

                if estado == Variables.Pendiente:
                    # Descargar el video o audio según el formato
                    descargar_t(enlace, formato, resolucion, nueva_carpeta)
                    # Incrementar Variables.fila para pasar a la siguiente fila en el próximo ciclo
                    Variables.fila += 1
                else: pass

            Msg = f"Se descargo la lista completa."
            Alertas.alerta_ok(Variables.titulo, "Aviso", Msg)
        except Exception as e:
            Msg = f"Ocurrió un error durante la descarga: {e}"
            Alertas.alerta_Amarilla(Variables.titulo, "Error imprevisto", Msg)
    else:
        Msg = "No se seleccionó ninguna carpeta de destino."
        Alertas.alerta_Amarilla(Variables.titulo, "Alerta", Msg)
        pass

def descargar_t(enlace, formato, resolucion, carpeta_destino):
    try:
        # Crear objeto YouTube para el video
        yt = YouTube(enlace)

        if formato == "Mp4":
            video_stream = yt.streams.filter(file_extension="mp4", progressive=True, res=resolucion).first()
            if video_stream:
                video_stream.download(output_path=carpeta_destino)
                actualizar_estado_lista(Variables.Terminado, enlace)
            else:
                Msg = f"La calidad seleccionada {resolucion} no está disponible para esta lista. Seleccione una menor e inténtelo de nuevo."
                Alertas.alerta_Amarilla(Variables.titulo, "Error calidad", Msg)
        else:
            # Descargar audio en formato MP3
            audio_stream = yt.streams.filter(only_audio=True).first()
            if audio_stream:
                audio_filename = audio_stream.default_filename
                audio_stream.download(output_path=carpeta_destino)
                # Cambiar la extensión del archivo a .mp3
                audio_mp4_path = os.path.join(carpeta_destino, audio_filename)
                audio_mp3_path = os.path.splitext(audio_mp4_path)[0] + ".mp3"
                # Renombrar el archivo con la extensión .mp3
                os.rename(audio_mp4_path, audio_mp3_path)
                actualizar_estado_lista(Variables.Terminado, enlace)
    except Exception as e:
        Msg = f"Ocurrió un error durante la descarga: {e}"
        Alertas.alerta_Amarilla(Variables.titulo, "Error imprevisto", Msg)
